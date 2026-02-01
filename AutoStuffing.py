"""
Autostuffing — сборка отгрузочных таблиц из папок со счетами.
Анализирует имена папок, группирует по приложению и товару, копирует листы из xlsx,
заполняет лист Total (счета, ЭСД, GTD).
"""

import re
import shutil
import subprocess
import sys
import time
from pathlib import Path
from collections import defaultdict
from copy import copy

# =============================================================================
# КОНФИГУРАЦИЯ — можно менять при масштабировании (другие папки, пути, имена)
# =============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent        # Папка скрипта (templates рядом)
REQUIREMENTS_FILE = SCRIPT_DIR / "requirements.txt"
INSTALLED_MARKER = SCRIPT_DIR / ".requirements_installed"  # Есть = библиотеки уже ставили (не ставить повторно)
INVOICE_SUBFOLDER = "invoices"                      # Подпапка со счетами (относительно введённой папки)
TEMPLATE_SUBFOLDER = "templates"                    # Папка с шаблоном (xlsx или xlsm) — всегда в папке скрипта
# Результаты сохраняются в рабочую папку (root), подпапка output не создаётся
LAST_PATH_FILE = SCRIPT_DIR / "path.txt"  # Файл с последней рабочей папкой
# -----------------------------------------------------------------------


# =============================================================================
# УСТАНОВКА ЗАВИСИМОСТЕЙ — при первом запуске pip install -r requirements.txt
# При масштабировании: сменить REQUIREMENTS_FILE / INSTALLED_MARKER
# =============================================================================

def _ensure_requirements_installed() -> None:
    """При первом запуске ставит библиотеки из requirements.txt; при следующих — ничего не делает."""
    if INSTALLED_MARKER.exists():
        return
    if not REQUIREMENTS_FILE.is_file():
        return
    try:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "-r", str(REQUIREMENTS_FILE)],
            check=False,
            capture_output=True,
        )
        INSTALLED_MARKER.write_text("", encoding="utf-8")
    except Exception:
        pass


_ensure_requirements_installed()

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Установите openpyxl: pip install openpyxl")

# =============================================================================
# ЦВЕТА И КОНСОЛЬ — ANSI-коды для вывода в терминал Windows
# При масштабировании: добавить свои COLORS, поменять _PATH_COLOR
# =============================================================================

COLORS = [
    "\033[92m",   # светло-зелёный
    "\033[96m",   # светло-голубой
    "\033[93m",   # светло-жёлтый
    "\033[95m",   # светло-пурпурный
    "\033[94m",   # светло-синий
    "\033[97m",   # белый
]
RESET = "\033[0m"


def _enable_ansi_windows() -> None:
    """Включает ANSI-цвета в консоли Windows."""
    try:
        if sys.platform == "win32":
            import ctypes
            k = ctypes.windll.kernel32
            h = k.GetStdHandle(-11)
            m = (k.GetConsoleMode(h) | 4) & 0xFFFF
            k.SetConsoleMode(h, m)
    except Exception:
        pass


def _save_last_path(path: Path) -> None:
    try:
        LAST_PATH_FILE.write_text(str(path.resolve()), encoding="utf-8")
    except Exception:
        pass


def _load_last_path() -> Path | None:
    try:
        if LAST_PATH_FILE.is_file():
            p = Path(LAST_PATH_FILE.read_text(encoding="utf-8").strip())
            if p.is_dir():
                return p
    except Exception:
        pass
    return None


_PATH_COLOR = "\033[93m"  # Жёлтый для текущего пути


# =============================================================================
# РАБОЧАЯ ПАПКА — запрос пути, сохранение в path.txt
# При масштабировании: сменить LAST_PATH_FILE, логику ask_work_directory
# =============================================================================

def ask_work_directory() -> Path:
    """
    Запрашивает рабочую папку. Если путь уже вводился — предлагает Enter (оставить) или новый путь.
    Текущий путь выводится жёлтым цветом.
    """
    last = _load_last_path()
    if last is not None:
        prompt = f"  Рабочая папка:\n  {_PATH_COLOR}{last}{RESET}\n  {COLORS[0]}Enter — тот же путь{RESET}  или  вставьте новый путь: "
    else:
        prompt = "  Укажите папку для работы (путь к папке со счетами): "
    raw = input(prompt).strip()
    if last is not None and raw == "":
        return last
    if (raw.startswith('"') and raw.endswith('"')) or (raw.startswith("'") and raw.endswith("'")):
        raw = raw[1:-1]
    path = Path(raw).resolve()
    if not path.is_dir():
        raise FileNotFoundError(f"Папка не найдена: {path}")
    _save_last_path(path)
    return path


# =============================================================================
# ГРУППИРОВКА ПАПОК — приложение (после 3-й запятой), порода/товар (после 5-й запятой)
# "N pack" игнорируется; "spr 16 pack" → "spr"
# При масштабировании: поменять get_group_key (части 3, 4, 5), паттерны _PACK_*
# =============================================================================

def get_application_name(folder_name: str) -> str:
    """Имя приложения = 4-я часть (после 3-й запятой)."""
    parts = parse_name_by_commas(folder_name)
    return parts[3] if len(parts) > 3 else folder_name


_PACK_PATTERN = re.compile(r"^\d+\s*pack$", re.IGNORECASE)
# Убирает " N pack" с конца (spr 16 pack → spr)
_PACK_SUFFIX_PATTERN = re.compile(r"\s+\d+\s*pack\s*$", re.IGNORECASE)


def _strip_pack_suffix(s: str) -> str:
    """Убирает ' N pack' с конца строки (spr 16 pack → spr)."""
    if not s:
        return s
    return _PACK_SUFFIX_PATTERN.sub("", s).strip()


def get_group_key(folder_name: str) -> str:
    """
    Ключ группировки: приложение (после 3-й запятой) + часть 4 + порода/товар (после 5-й запятой).
    "spr 16 pack" → "spr"; не создавать отдельную таблицу под "N pack".
    """
    parts = parse_name_by_commas(folder_name)
    if len(parts) < 3:
        return ""
    app = parts[3] if len(parts) > 3 else folder_name
    key_parts = [app]
    if len(parts) > 4:
        key_parts.append(parts[4].strip())
    if len(parts) > 5:
        product = _strip_pack_suffix(parts[5].strip())
        if product and not _PACK_PATTERN.match(product):
            key_parts.append(product)
    return " | ".join(key_parts)


def parse_name_by_commas(name: str) -> list[str]:
    """Разбить имя по запятым (1-я, 2-я, 3-я части и т.д.)."""
    return [p.strip() for p in name.split(",")]


def _invoice_numbers_to_range_string(numbers: list[str]) -> str:
    """
    Преобразует список номеров счетов в строку диапазонов: (43;93;95-97;100).
    Подряд идущие числа объединяются в диапазон через дефис, остальные через точку с запятой.
    """
    nums: list[int] = []
    for n in numbers:
        s = str(n).strip()
        if s.isdigit():
            nums.append(int(s))
    nums.sort()
    if not nums:
        return "()"
    parts: list[str] = []
    i = 0
    while i < len(nums):
        start = nums[i]
        j = i + 1
        while j < len(nums) and nums[j] == nums[j - 1] + 1:
            j += 1
        end = nums[j - 1]
        if start == end:
            parts.append(str(start))
        else:
            parts.append(f"{start}-{end}")
        i = j
    return "(" + ";".join(parts) + ")"


# =============================================================================
# ИМЯ ОТГРУЗОЧНОЙ ТАБЛИЦЫ — диапазон счетов, части из папки; "N pack" убирается
# При масштабировании: parse_name_by_commas, _invoice_numbers_to_range_string, формат имени
# =============================================================================

def build_upload_table_filename(
    template_name: str,
    invoice_folder_name: str,
    invoice_numbers: list[str] | None = None,
) -> str:
    """
    Имя отгрузочной таблицы. Если передан invoice_numbers — в начало ставится
    диапазон и количество: (43;93;95-97;100) 6 pcs., иначе первая часть из шаблона (напр. () pcs.).
    Остальное: вторая часть из шаблона + части из папки со счетами (после 2-й запятой).
    Пример: "(43;93;95-97;100) 6 pcs., LI, 40_2023, Add. VP-CH-2510-23, ZLPK (TS), pine"
    """
    t_parts = parse_name_by_commas(template_name)
    # "N pack" не попадает в имя; "spr 16 pack" → "spr"
    i_parts = []
    for p in parse_name_by_commas(invoice_folder_name):
        p = p.strip()
        if _PACK_PATTERN.match(p):  # отдельная часть "13 pack" — пропускаем
            continue
        p = _strip_pack_suffix(p)  # "spr 16 pack" → "spr"
        if p:
            i_parts.append(p)
    if invoice_numbers:
        range_str = _invoice_numbers_to_range_string(invoice_numbers)
        count = len(invoice_numbers)
        first_part = f"{range_str} {count} pcs."
    else:
        first_part = t_parts[0] if t_parts else "() pcs."
    # Первая часть — диапазон+кол-во или из шаблона, вторая из шаблона, остальное из папки
    result = [first_part] + (t_parts[1:2] if len(t_parts) > 1 else []) + i_parts[2:]
    name_parts = [p for p in result if p]
    raw = ", ".join(name_parts)
    return re.sub(r'[\\/:*?"<>|]', "_", raw).strip(",_") or "upload_table"


# =============================================================================
# КОПИРОВАНИЕ ЛИСТОВ — ячейки, формат, ширина столбцов, область печати
# Все столбцы/строки видимы, масштаб 25%
# При масштабировании: _SHEET_ZOOM_PERCENT, _copy_sheet_print_and_view, _copy_cell_style
# =============================================================================

def _copy_cell_style(src_cell, tgt_cell) -> None:
    """Копирует значение и оформление ячейки (между книгами). Формулы копируются как формулы."""
    tgt_cell.value = src_cell.value
    if src_cell.has_style:
        if src_cell.font:
            tgt_cell.font = copy(src_cell.font)
        if src_cell.border:
            tgt_cell.border = copy(src_cell.border)
        if src_cell.fill:
            tgt_cell.fill = copy(src_cell.fill)
        if src_cell.alignment:
            tgt_cell.alignment = copy(src_cell.alignment)
        if src_cell.number_format:
            tgt_cell.number_format = src_cell.number_format


def _copy_sheet_print_and_view(ws_src, ws_tgt) -> None:
    """Копирует область печати, настройки страницы и видимую область (как при «Копировать» в Excel)."""
    try:
        if getattr(ws_src, "print_area", None):
            ws_tgt.print_area = ws_src.print_area
        if getattr(ws_src, "print_title_cols", None):
            ws_tgt.print_title_cols = ws_src.print_title_cols
        if getattr(ws_src, "print_title_rows", None):
            ws_tgt.print_title_rows = ws_src.print_title_rows
        if getattr(ws_src, "page_margins", None) and ws_src.page_margins is not None:
            ws_tgt.page_margins = copy(ws_src.page_margins)
        if getattr(ws_src, "page_setup", None) and ws_src.page_setup is not None:
            ws_tgt.page_setup = copy(ws_src.page_setup)
        if getattr(ws_src, "print_options", None) and ws_src.print_options is not None:
            ws_tgt.print_options = copy(ws_src.print_options)
        # views не копируем — там может быть скрытие столбцов/строк
        if getattr(ws_src, "freeze_panes", None) is not None:
            ws_tgt.freeze_panes = ws_src.freeze_panes
        # Масштаб листа — 25%, чтобы весь лист был виден
        _set_sheet_zoom_25(ws_tgt)
    except Exception as e:
        print(f"  [ошибка] настройки листа: {e}")


# Масштаб по умолчанию для скопированных листов (весь лист виден)
_SHEET_ZOOM_PERCENT = 25


def _set_sheet_zoom_25(ws_tgt) -> None:
    """Устанавливает масштаб 25% на листе, чтобы весь лист был виден."""
    try:
        sv = getattr(ws_tgt, "sheet_view", None)
        if sv is None:
            return
        sv.zoomScale = _SHEET_ZOOM_PERCENT
        if hasattr(sv, "zoomScaleNormal"):
            sv.zoomScaleNormal = _SHEET_ZOOM_PERCENT
    except Exception as e:
        print(f"  [ошибка] масштаб листа: {e}")


def get_first_sheet_name(xlsx_path: Path) -> str:
    """Возвращает имя первого листа в xlsx (номер/имя счёта для проверки «уже в отгрузочной таблице»)."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        return wb.worksheets[0].title
    finally:
        wb.close()


def copy_first_sheet_to_workbook(
    source_xlsx_path: Path,
    target_wb,
    new_sheet_name: str | None = None,
) -> None:
    """Копирует первый лист из xlsx в целевую книгу (отгрузочную таблицу): ячейки, формат, область печати, настройки страницы, видимую область."""
    wb_src = load_workbook(source_xlsx_path, read_only=False, data_only=False)
    ws_src = wb_src.worksheets[0]
    name = (new_sheet_name or ws_src.title)[:31]
    existing = {s.title for s in target_wb.worksheets}
    if name in existing:
        base, n = name, 1
        while f"{base}_{n}" in existing:
            n += 1
        name = f"{base}_{n}"[:31]
    ws_tgt = target_wb.create_sheet(title=name)
    for row in range(1, ws_src.max_row + 1):
        for col in range(1, ws_src.max_column + 1):
            src_c = ws_src.cell(row=row, column=col)
            tgt_c = ws_tgt.cell(row=row, column=col)
            _copy_cell_style(src_c, tgt_c)
    # Ширина столбцов — копируем размеры, скрытие НЕ копируем (все столбцы видимы)
    for col_key, src_dim in list(ws_src.column_dimensions.items()):
        if src_dim is None or getattr(src_dim, "width", None) is None:
            continue
        w = src_dim.width
        min_col = getattr(src_dim, "min", None)
        max_col = getattr(src_dim, "max", None)
        if min_col is not None and max_col is not None:
            for c in range(min_col, max_col + 1):
                col_letter = get_column_letter(c)
                ws_tgt.column_dimensions[col_letter].width = w
                ws_tgt.column_dimensions[col_letter].hidden = False
        else:
            col_letter = get_column_letter(col_key) if isinstance(col_key, int) else str(col_key)
            ws_tgt.column_dimensions[col_letter].width = w
            ws_tgt.column_dimensions[col_letter].hidden = False
    # Высота строк — копируем размеры, скрытие НЕ копируем (все строки видимы)
    for row_key, src_dim in list(ws_src.row_dimensions.items()):
        if src_dim is not None and getattr(src_dim, "height", None) is not None:
            ws_tgt.row_dimensions[row_key].height = src_dim.height
            ws_tgt.row_dimensions[row_key].hidden = False
    for merged in ws_src.merged_cells.ranges:
        ws_tgt.merge_cells(str(merged))
    # Снимаем скрытие со всех столбцов и строк — как в счёте, но всё видно
    for c in range(1, ws_src.max_column + 1):
        ws_tgt.column_dimensions[get_column_letter(c)].hidden = False
    for row_key in list(ws_tgt.row_dimensions):
        ws_tgt.row_dimensions[row_key].hidden = False
    _copy_sheet_print_and_view(ws_src, ws_tgt)
    # Режим просмотра «Страничный режим» (Page Break Preview) + масштаб 25%
    try:
        ws_tgt.sheet_view.view = "pageBreakPreview"
        _set_sheet_zoom_25(ws_tgt)
    except Exception:
        pass
    wb_src.close()


# =============================================================================
# ЭСД И GTD — PDF в папках счетов: ЭСД — любой PDF, GTD — GTD_a_b_c.pdf
# При масштабировании: _ESD_PATTERN, _GTD_PATTERN, _COL_ESD, _COL_GTD, _SKIP_SHEET_TITLES
# =============================================================================

_ESD_PATTERN = re.compile(r"^([\w-]+)\.pdf$", re.IGNORECASE)
_GTD_PATTERN = re.compile(r"^GTD_(\d+)_(\d+)_(\d+)\.pdf$", re.IGNORECASE)
# Колонки на листе Total: J = ЭСД, O = декларации (GTD)
_COL_ESD = 10   # J
_COL_GTD = 15   # O
# Прочерк в ячейке, если в папке счёта нет ЭСД или GTD
_EMPTY_PASS = "—"
# Имена листов, не считающиеся счетами (лист Total и т.п.)
_SKIP_SHEET_TITLES = ("TOTAL", "CONSOLIDATED INVOICE")


def _find_total_sheet(wb):
    """Возвращает лист «Total» (без учёта регистра) или None."""
    for ws in wb.worksheets:
        if ws.title.strip().upper() == "TOTAL":
            return ws
    return None


def _get_sorted_invoice_numbers_from_wb(wb) -> list[str]:
    """Список номеров счетов из имён листов книги (кроме первого, Total, Consolidated Invoice), отсортированный по числу."""
    names = [
        s.title for s in wb.worksheets[1:]
        if s.title.strip().upper() not in _SKIP_SHEET_TITLES
    ]
    return _sort_invoice_numbers_as_int(names) if names else []


def _collect_esd_and_gtd_from_one_folder(folder: Path) -> tuple[list[str], list[str]]:
    """
    Сканирует одну папку счёта: номера ЭСД (PDF не GTD_) и номера деклараций (GTD_a_b_c.pdf → a/b/c).
    Возвращает (esd_list, gtd_list).
    """
    esd_list: list[str] = []
    gtd_list: list[str] = []
    if not folder.is_dir():
        return esd_list, gtd_list
    for f in folder.iterdir():
        if not f.is_file():
            continue
        name = f.name
        if name.startswith("GTD_"):
            gtd_match = _GTD_PATTERN.match(name)
            if gtd_match:
                gtd_list.append(f"{gtd_match.group(1)}/{gtd_match.group(2)}/{gtd_match.group(3)}")
            continue
        esd_match = _ESD_PATTERN.match(name)
        if esd_match:
            esd_list.append(name[:-4])
    return esd_list, gtd_list


def _fill_total_sheet(wb, inv_to_folder: dict[str, Path]) -> None:
    """
    Один проход по листу «Total»: заполняет B4.. номерами счетов, J4.. — ЭСД, O4.. — декларации (GTD).
    Если в папке счёта нет ЭСД или GTD — в ячейке прочерк «—».
    """
    try:
        target = _find_total_sheet(wb)
        if target is None:
            return
        sorted_numbers = _get_sorted_invoice_numbers_from_wb(wb)
        if not sorted_numbers:
            target.cell(row=4, column=2, value="")
            return
        n = len(sorted_numbers)
        for r in range(4, 4 + n):
            target.cell(row=r, column=2, value="")
            target.cell(row=r, column=_COL_ESD, value="")
            target.cell(row=r, column=_COL_GTD, value="")
        for i, inv_num in enumerate(sorted_numbers):
            target.cell(row=4 + i, column=2, value=inv_num)
            folder = inv_to_folder.get(inv_num)
            if folder is None:
                target.cell(row=4 + i, column=_COL_ESD, value=_EMPTY_PASS)
                target.cell(row=4 + i, column=_COL_GTD, value=_EMPTY_PASS)
            else:
                esd_list, gtd_list = _collect_esd_and_gtd_from_one_folder(folder)
                esd_str = ", ".join(esd_list) if esd_list else _EMPTY_PASS
                gtd_str = ", ".join(gtd_list) if gtd_list else _EMPTY_PASS
                target.cell(row=4 + i, column=_COL_ESD, value=esd_str)
                target.cell(row=4 + i, column=_COL_GTD, value=gtd_str)
    except Exception as e:
        print(f"  [ошибка] лист Total: {e}")


def _invoice_number_sort_key(s: str) -> int:
    """Ключ для сортировки по числовому значению номера счёта (1, 2, 3 … 10, 11)."""
    s = str(s).strip()
    if s.isdigit():
        return int(s)
    return 0


def _sort_invoice_numbers_as_int(numbers: list[str]) -> list[str]:
    """Сортирует номера счетов по возрастанию числа (2, 10, 100, 110, 2511)."""
    return sorted(numbers, key=_invoice_number_sort_key)


def count_invoices_in_folders(folders: list[Path]) -> int:
    """Количество xlsx-счетов в переданных папках."""
    return sum(
        1 for folder in folders
        for p in folder.rglob("*.xlsx")
        if not p.name.startswith("~$")
    )


def _folders_without_xlsx(folders: list[Path]) -> list[Path]:
    """Папки, в которых нет ни одного .xlsx (кроме ~$)."""
    empty: list[Path] = []
    for folder in folders:
        if not folder.is_dir():
            continue
        has_xlsx = any(not p.name.startswith("~$") for p in folder.rglob("*.xlsx"))
        if not has_xlsx:
            empty.append(folder)
    return empty


def _count_esd_gtd_in_folders(folders: list[Path]) -> tuple[int, int]:
    """Количество документов ЭСД и GTD (PDF) в папках. Возвращает (число_ЭСД, число_GTD)."""
    esd_count, gtd_count = 0, 0
    for folder in folders:
        if not folder.is_dir():
            continue
        for f in folder.iterdir():
            if not f.is_file():
                continue
            name = f.name
            if name.startswith("GTD_"):
                if _GTD_PATTERN.match(name):
                    gtd_count += 1
                continue
            if _ESD_PATTERN.match(name):
                esd_count += 1
    return esd_count, gtd_count


# =============================================================================
# АНАЛИЗ ПАПОК — сканирование, группировка по get_group_key
# При масштабировании: мин. число частей (3), get_group_key
# =============================================================================

def analyze_and_group_invoice_folders(base_path: Path) -> dict[str, list[Path]]:
    """
    Сканирует папки со счетами, группирует по приложению + тип (приложение = часть после 3-й запятой,
    тип = последняя часть). Пример: "..., WPCH 0906, Demand" и "..., WPCH 0906, PORODA Pine" — разные таблицы.
    Учитываются только папки с не менее чем 3 частями в имени.
    """
    if not base_path.is_dir():
        return {}
    by_group = defaultdict(list)
    for item in base_path.iterdir():
        if item.is_dir() and not item.name.startswith("."):
            parts = parse_name_by_commas(item.name)
            if len(parts) < 3:
                continue
            key = get_group_key(item.name)
            if not key:
                continue
            by_group[key].append(item)
    return dict(by_group)


# =============================================================================
# ОБРАБОТКА ГРУППЫ — создание/обновление отгрузочной таблицы, копирование листов
# При масштабировании: build_upload_table_filename, copy_first_sheet_to_workbook
# =============================================================================

def process_application(
    app_name: str,
    invoice_folders: list[Path],
    template_path: Path,
    output_dir: Path,
) -> int:
    """
    Для каждого приложения: если отгрузочная таблица уже есть — добавляем только новые счета
    (по имени первого листа: если такой лист уже в таблице, счёт не копируем). Если таблицы нет —
    создаём из шаблона и копируем первый лист из всех xlsx.
    """
    template_name = template_path.stem
    first_folder_name = invoice_folders[0].name
    ext = template_path.suffix.lower()
    keep_vba = ext == ".xlsm"

    # Папки без ни одного .xlsx — сообщаем пользователю
    for folder in _folders_without_xlsx(invoice_folders):
        print(f"  {COLORS[2]}Папка без счетов (xlsx): {folder.name}{RESET}")

    file_base = build_upload_table_filename(template_name, first_folder_name)
    out_path = output_dir / f"{file_base}{ext}"

    if out_path.exists():
        wb = load_workbook(out_path, keep_vba=keep_vba)
        existing_sheet_names = {s.title for s in wb.worksheets}
        inv_to_folder: dict[str, Path] = {}
        new_paths: list[tuple[Path, str]] = []
        for folder in invoice_folders:
            if not folder.is_dir():
                continue
            for path in folder.rglob("*.xlsx"):
                if path.name.startswith("~$"):
                    continue
                try:
                    sheet_name = get_first_sheet_name(path)
                    inv_to_folder[sheet_name] = folder
                    if sheet_name not in existing_sheet_names:
                        new_paths.append((path, sheet_name))
                except Exception as e:
                    print(f"  Пропуск {path.name}: {e}")
        new_paths.sort(key=lambda x: _invoice_number_sort_key(x[1]))
        added = 0
        if new_paths:
            for path, _ in new_paths:
                try:
                    copy_first_sheet_to_workbook(path, wb)
                    added += 1
                    name = wb.worksheets[-1].title
                    print(f"  {COLORS[0]}{name} скопирован{RESET}")
                except Exception as e:
                    print(f"  Пропуск {path.name}: {e}")
        _fill_total_sheet(wb, inv_to_folder)
        wb.save(out_path)
        wb.close()
        if added:
            print(f"  {COLORS[1]}Готово: {out_path.name} (добавлено листов: {added}){RESET}")
        else:
            print(f"  {COLORS[0]}Готово: {out_path.name} (обновлены ЭСД/GTD на листе Total){RESET}")
        return added, out_path
    else:
        all_paths: list[tuple[Path, Path]] = []
        for folder in invoice_folders:
            if not folder.is_dir():
                continue
            for path in folder.rglob("*.xlsx"):
                if path.name.startswith("~$"):
                    continue
                all_paths.append((path, folder))
        # Сортируем по числовому значению номера счёта (1, 2, 3 … 10, 11)
        all_with_name: list[tuple[Path, Path, str]] = []
        for path, folder in all_paths:
            try:
                sn = get_first_sheet_name(path)
                all_with_name.append((path, folder, sn))
            except Exception:
                pass
        all_with_name.sort(key=lambda x: _invoice_number_sort_key(x[2]))
        file_base = build_upload_table_filename(template_name, first_folder_name)
        out_path = output_dir / f"{file_base}{ext}"
        shutil.copy2(template_path, out_path)
        wb = load_workbook(out_path, keep_vba=keep_vba)
        inv_to_folder: dict[str, Path] = {}
        copied = 0
        for path, folder, _ in all_with_name:
            try:
                copy_first_sheet_to_workbook(path, wb)
                inv_to_folder[wb.worksheets[-1].title] = folder
                copied += 1
                name = wb.worksheets[-1].title
                print(f"  {COLORS[0]}{name} скопирован{RESET}")
            except Exception as e:
                print(f"  Пропуск {path.name}: {e}")
        _fill_total_sheet(wb, inv_to_folder)
        wb.save(out_path)
        wb.close()
        print(f"  {COLORS[1]}Готово: {out_path}{RESET}")
        return copied, out_path


# =============================================================================
# ГЛАВНЫЙ ЦИКЛ — запрос пути, поиск шаблона, обработка групп, итоги, выбор 1/2
# При масштабировании: структура main, выбор переименования, вывод в консоль
# =============================================================================

def main():
    _enable_ansi_windows()
    print("\n  === Отгрузочные таблицы ===")
    print("  Укажите путь или Enter (тот же путь).\n")
    try:
        root = ask_work_directory()
    except FileNotFoundError as e:
        print(e)
        return

    template_folder = SCRIPT_DIR / TEMPLATE_SUBFOLDER
    if not template_folder.is_dir():
        print(f"Папка шаблона не найдена: {template_folder}")
        print("Папка 'templates' должна быть в одной папке со скриптом Autostuffing.py")
        return
    # Шаблон — любой .xlsx или .xlsm в папке templates (без временных ~$)
    template_files = [
        f for f in template_folder.iterdir()
        if f.is_file() and f.suffix.lower() in (".xlsx", ".xlsm") and not f.name.startswith("~$")
    ]
    if not template_files:
        print(f"В папке {template_folder} нет файла .xlsx или .xlsm. Добавьте туда отгрузочную таблицу.")
        return
    template_path = template_files[0]
    if len(template_files) > 1:
        print(f"Используется шаблон: {template_path.name}")

    # 1) Analyze: get invoice folders (direct children of invoices root)
    invoices_root = root / INVOICE_SUBFOLDER
    if not invoices_root.is_dir():
        invoices_root = root
    by_app = analyze_and_group_invoice_folders(invoices_root)

    if not by_app:
        print("Папки со счетами не найдены:", invoices_root)
        return

    output_dir = root

    # Обработка приложений; список файлов для опционального переименования (диапазон в имя)
    total_processed = 0
    rename_list: list[tuple[Path, str, str]] = []  # (out_path, template_name, first_folder_name)
    for i, (app_name, folders) in enumerate(by_app.items()):
        app_color = COLORS[i % len(COLORS)]
        print(f"\n  {app_color}─── Приложение: {app_name} ───{RESET}")
        try:
            count, out_path = process_application(
                app_name,
                folders,
                template_path,
                output_dir,
            )
            total_processed += count
            rename_list.append((out_path, template_path.stem, folders[0].name))
        except Exception as e:
            print(f"  Ошибка: {e}")

    # Итоговая сводка: приложения, счета, ЭСД, GTD
    print("\n  " + "=" * 52)
    print("  ИТОГИ")
    print("  " + "=" * 52)
    total_inv, total_esd, total_gtd = 0, 0, 0
    for i, (app_name, folders) in enumerate(by_app.items()):
        inv = count_invoices_in_folders(folders)
        esd, gtd = _count_esd_gtd_in_folders(folders)
        total_inv += inv
        total_esd += esd
        total_gtd += gtd
        color = COLORS[i % len(COLORS)]
        print(f"  {color}{app_name}{RESET}")
        print(f"      счетов: {inv:>4}   ЭСД: {esd:>4}   GTD: {gtd:>4}")
    print("  " + "-" * 52)
    print(f"  Всего приложений: {len(by_app)}")
    print(f"  Всего инвойсов: {total_inv:>4}")
    print(f"  Всего ЭСД: {total_esd:>4}")
    print(f"  Всего ДТ: {total_gtd:>4}")
    print("  " + "=" * 52)

    # Выбор: закрыть без изменения имён или добавить диапазон счетов в имена
    while True:
        print(f"\n  {COLORS[4]}1 — закрыть без изменения имён файлов{RESET}")
        print(f"  {COLORS[4]}2 — добавить диапазон счетов и количество в имена файлов{RESET}")
        choice = input("  Ваш выбор (1 или 2): ").strip()
        if choice == "1":
            break
        if choice == "2":
            ext = template_path.suffix.lower()
            keep_vba = ext == ".xlsm"
            for out_path, template_name, first_folder_name in rename_list:
                if not out_path.exists():
                    continue
                try:
                    wb = load_workbook(out_path, keep_vba=keep_vba)
                    numbers = _get_sorted_invoice_numbers_from_wb(wb)
                    wb.close()
                    long_base = build_upload_table_filename(
                        template_name, first_folder_name, numbers
                    )
                    new_path = output_dir / f"{long_base}{ext}"
                    if new_path.resolve() == out_path.resolve():
                        continue
                    out_path.rename(new_path)
                    print(f"  {COLORS[1]}Переименовано: {out_path.name} → {new_path.name}{RESET}")
                except Exception as e:
                    print(f"  {COLORS[2]}[ошибка] {out_path.name}: {e}{RESET}")
            break
        print("  Введите 1 или 2.")

    return total_processed


def main_return_none_on_error() -> int | None:
    """Возвращает число обработанных файлов или None при ошибке."""
    try:
        return main()
    except Exception as e:
        print("Ошибка:", e)
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    total = main_return_none_on_error()
    if total is not None:
        print()
        if total == 0:
            print(f"  {COLORS[0]}Новых листов не добавлено (таблицы обновлены).{RESET}")
        else:
            print(f"  {COLORS[1]}Добавлено листов в таблицы: {total}{RESET}")
    print(f"\n  {COLORS[4]}Нажмите Enter для выхода...{RESET}")
    input()
